# pip install transformers torch Pillow googletrans==4.0.0-rc1 openpyxl

import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk, ImageDraw, ImageFont
import threading
import os

from transformers import BlipProcessor, BlipForConditionalGeneration, DetrImageProcessor, DetrForObjectDetection
import torch
from googletrans import Translator
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

CONFIDENCE_THRESHOLD = 0.8

blip_processor, blip_model, detr_processor, detr_model = None, None, None, None
models_loaded = threading.Event()

COCO_LABELS_FR = {
    'person': 'personne', 'bicycle': 'vélo', 'car': 'voiture', 'motorcycle': 'moto',
    'airplane': 'avion', 'bus': 'bus', 'train': 'train', 'truck': 'camion', 'boat': 'bateau',
    'traffic light': 'feu de circulation', 'fire hydrant': 'bouche d\'incendie', 'stop sign': 'panneau stop',
    'parking meter': 'parcmètre', 'bench': 'banc', 'bird': 'oiseau', 'cat': 'chat', 'dog': 'chien',
    'horse': 'cheval', 'sheep': 'mouton', 'cow': 'vache', 'elephant': 'éléphant', 'bear': 'ours',
    'zebra': 'zèbre', 'giraffe': 'girafe', 'backpack': 'sac à dos', 'umbrella': 'parapluie',
    'handbag': 'sac à main', 'tie': 'cravate', 'suitcase': 'valise', 'frisbee': 'frisbee',
    'skis': 'skis', 'snowboard': 'snowboard', 'sports ball': 'ballon de sport', 'kite': 'cerf-volant',
    'baseball bat': 'batte de baseball', 'baseball glove': 'gant de baseball', 'skateboard': 'skateboard',
    'surfboard': 'planche de surf', 'tennis racket': 'raquette de tennis', 'bottle': 'bouteille',
    'wine glass': 'verre à vin', 'cup': 'tasse', 'fork': 'fourchette', 'knife': 'couteau',
    'spoon': 'cuillère', 'bowl': 'bol', 'banana': 'banane', 'apple': 'pomme', 'sandwich': 'sandwich',
    'orange': 'orange', 'broccoli': 'brocoli', 'carrot': 'carotte', 'hot dog': 'hot-dog',
    'pizza': 'pizza', 'donut': 'donut', 'cake': 'gâteau', 'chair': 'chaise', 'couch': 'canapé',
    'potted plant': 'plante en pot', 'bed': 'lit', 'dining table': 'table à manger', 'toilet': 'toilettes',
    'tv': 'télévision', 'laptop': 'ordinateur portable', 'mouse': 'souris', 'remote': 'télécommande',
    'keyboard': 'clavier', 'cell phone': 'téléphone portable', 'microwave': 'micro-ondes',
    'oven': 'four', 'toaster': 'grille-pain', 'sink': 'évier', 'refrigerator': 'réfrigérateur',
    'book': 'livre', 'clock': 'horloge', 'vase': 'vase', 'scissors': 'ciseaux', 'teddy bear': 'ours en peluche',
    'hair drier': 'sèche-cheveux', 'toothbrush': 'brosse à dents'
}

class RoundedFrame(tk.Frame):
    def __init__(self, parent, radius=20, bg_color=None, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self.radius = radius; self.bg_color = bg_color if bg_color else parent["bg"]; self.corner_image = None
        self.bind("<Configure>", self._draw_rounded_corners)
    def _draw_rounded_corners(self, event=None):
        width, height = self.winfo_width(), self.winfo_height()
        if width < 2 * self.radius or height < 2 * self.radius: return
        if self.corner_image is None or self.corner_image.size != (width, height):
            self.corner_image = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(self.corner_image)
            draw.rounded_rectangle((0, 0, width, height), self.radius, fill=self.bg_color)
            self._photoimage = ImageTk.PhotoImage(self.corner_image)
            if hasattr(self, 'bg_label'): self.bg_label.config(image=self._photoimage)
            else: self.bg_label = tk.Label(self, image=self._photoimage, bd=0, bg=self["bg"]); self.bg_label.place(x=0, y=0, relwidth=1, relheight=1); self.bg_label.lower()

class FullImageAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Analyseur d'Image Avancé")
        self.root.geometry("1400x900"); self.root.minsize(1000, 700); self.root.configure(bg="#1e1e1e")
        self.BG_COLOR = "#1e1e1e"; self.CANVAS_BG = "#2a2a2a"; self.PANEL_BG = "#252526"; self.ITEM_BG = "#3c3c3c"
        self.TEXT_COLOR = "#d4d4d4"; self.ACCENT_COLOR = "#007acc"; self.HIGHLIGHT_COLOR = "#e45252"; self.FONT_FAMILY = "Segoe UI"
        self.object_widgets = []; self.last_results = None; self.original_image = None; self.last_description = ""
        self.displayed_image_pil = None; self.displayed_image_tk = None; self.image_on_canvas = None; self.current_hovered_index = None
        self.translator = Translator()
        self.create_widgets()
        self.show_welcome_screen()

    def create_widgets(self):
        self.top_bar = tk.Frame(self.root, bg=self.BG_COLOR)
        self.btn_reupload_top = tk.Button(self.top_bar, text="🖼️ Importer une autre image", command=self.start_analysis_thread, font=(self.FONT_FAMILY, 12), bg=self.ITEM_BG, fg=self.TEXT_COLOR, relief=tk.FLAT, padx=15, pady=8, activebackground=self.ACCENT_COLOR)
        self.btn_reupload_top.pack(pady=10); self.btn_reupload_top.bind("<Enter>", lambda e: e.widget.config(bg=self.ACCENT_COLOR)); self.btn_reupload_top.bind("<Leave>", lambda e: e.widget.config(bg=self.ITEM_BG))

        self.main_frame = tk.Frame(self.root, bg=self.CANVAS_BG); self.main_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); self.main_frame.bind("<Configure>", self.on_resize)
        self.canvas = tk.Canvas(self.main_frame, bg=self.CANVAS_BG, highlightthickness=0); self.canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.canvas.bind("<Motion>", self.on_canvas_hover); self.canvas.bind("<Leave>", self.on_canvas_leave)

        self.welcome_frame = tk.Frame(self.canvas, bg=self.CANVAS_BG)
        tk.Label(self.welcome_frame, text="🖼️", font=(self.FONT_FAMILY, 80), bg=self.CANVAS_BG, fg=self.ACCENT_COLOR).pack(pady=(0, 20))
        tk.Label(self.welcome_frame, text="Analyseur d'Image Avancé", font=(self.FONT_FAMILY, 24, "bold"), bg=self.CANVAS_BG, fg=self.TEXT_COLOR).pack(pady=10)
        self.btn_upload = tk.Button(self.welcome_frame, text="Commencer par Importer une Image", command=self.start_analysis_thread, font=(self.FONT_FAMILY, 14, "bold"), bg=self.ACCENT_COLOR, fg="white", relief=tk.FLAT, padx=20, pady=10, activebackground="#005a9e")
        self.btn_upload.pack(pady=20)
        self.loading_label = tk.Label(self.welcome_frame, text="Chargement...", font=(self.FONT_FAMILY, 16), bg=self.CANVAS_BG, fg=self.TEXT_COLOR)

        self.results_panel = RoundedFrame(self.root, radius=20, bg_color=self.PANEL_BG)
        tk.Label(self.results_panel, text="Résultats de l'Analyse", font=(self.FONT_FAMILY, 18, "bold"), bg=self.PANEL_BG, fg=self.ACCENT_COLOR).pack(pady=(20, 10), padx=25, anchor="w")
        tk.Label(self.results_panel, text="Description de la Scène", font=(self.FONT_FAMILY, 14, "bold"), bg=self.PANEL_BG, fg=self.TEXT_COLOR).pack(pady=(10,5), anchor="w", padx=25)
        self.description_text = tk.Label(self.results_panel, text="", font=(self.FONT_FAMILY, 12), wraplength=380, justify=tk.LEFT, bg=self.PANEL_BG, fg=self.TEXT_COLOR, anchor="w")
        self.description_text.pack(pady=5, padx=25, fill=tk.X)
        separator = tk.Frame(self.results_panel, height=2, bg=self.CANVAS_BG); separator.pack(fill=tk.X, padx=25, pady=15)
        tk.Label(self.results_panel, text="Objets Détectés", font=(self.FONT_FAMILY, 14, "bold"), bg=self.PANEL_BG, fg=self.TEXT_COLOR).pack(pady=(5,5), anchor="w", padx=25)
        self.objects_list_frame = tk.Frame(self.results_panel, bg=self.PANEL_BG); self.objects_list_frame.pack(pady=5, padx=25, fill=tk.BOTH, expand=True)
        
        action_frame = tk.Frame(self.results_panel, bg=self.PANEL_BG)
        action_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=15, padx=25)
        self.btn_save_excel = tk.Button(action_frame, text="💾 Sauvegarder en Excel", command=self.save_results_to_excel, font=(self.FONT_FAMILY, 12), bg=self.ITEM_BG, fg=self.TEXT_COLOR, relief=tk.FLAT, padx=15, pady=8)
        self.btn_save_excel.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        self.btn_reupload_panel = tk.Button(action_frame, text="🔄 Analyser une autre image", command=self.start_analysis_thread, font=(self.FONT_FAMILY, 12, "bold"), bg=self.ACCENT_COLOR, fg="white", relief=tk.FLAT, padx=15, pady=8)
        self.btn_reupload_panel.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(5, 0))
        self.btn_save_excel.bind("<Enter>", lambda e: e.widget.config(bg="#555")); self.btn_save_excel.bind("<Leave>", lambda e: e.widget.config(bg=self.ITEM_BG))
        self.btn_reupload_panel.bind("<Enter>", lambda e: e.widget.config(bg="#005a9e")); self.btn_reupload_panel.bind("<Leave>", lambda e: e.widget.config(bg=self.ACCENT_COLOR))
        
        self.status_label = tk.Label(self.root, text="Prêt", bd=1, relief=tk.FLAT, anchor=tk.W, bg="#1a1a1a", fg=self.TEXT_COLOR, padx=10); self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def load_models_in_thread(self):
        global blip_processor, blip_model, detr_processor, detr_model
        try:
            self.root.after(0, self.show_loading_screen, "Téléchargement des modèles IA...")
            blip_processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-large"); blip_model = BlipForConditionalGeneration.from_pretrained("Salesforce/blip-image-captioning-large")
            detr_processor = DetrImageProcessor.from_pretrained("facebook/detr-resnet-50"); detr_model = DetrForObjectDetection.from_pretrained("facebook/detr-resnet-50")
            models_loaded.set()
            self.root.after(0, self.update_status, "Modèles chargés. Prêt à analyser."); self.root.after(0, self.show_welcome_screen)
        except Exception as e: messagebox.showerror("Erreur de chargement", f"Impossible de charger les modèles IA.\n{e}"); self.root.quit()

    def start_analysis_thread(self):
        filepath = filedialog.askopenfilename(title="Choisissez une image", filetypes=[("Images", "*.jpg *.jpeg *.png")])
        if not filepath: return
        self.clear_previous_analysis()
        try: self.original_image = Image.open(filepath); threading.Thread(target=self.run_full_analysis, daemon=True).start()
        except Exception as e: messagebox.showerror("Erreur d'image", f"Impossible de charger ce fichier: {e}")

    def run_full_analysis(self):
        self.root.after(0, self.display_image); self.root.after(0, self.show_loading_screen, "Analyse en cours...")
        models_loaded.wait()
        desc_thread = threading.Thread(target=self.generate_description, args=(self.original_image,)); obj_thread = threading.Thread(target=self.detect_objects, args=(self.original_image,))
        desc_thread.start(); obj_thread.start(); desc_thread.join(); obj_thread.join()
        self.root.after(0, self.hide_welcome_screen); self.root.after(0, self.animate_results_panel_in)
        self.root.after(0, self.update_status, "Analyse terminée !"); self.root.after(0, self.top_bar.pack, {"side": "top", "fill": "x"})

    def generate_description(self, image):
        try:
            inputs = blip_processor(image.convert('RGB'), return_tensors="pt"); output = blip_model.generate(**inputs, max_length=150, num_beams=5)
            desc_en = blip_processor.decode(output[0], skip_special_tokens=True)
            try: desc_fr = self.translator.translate(desc_en, src='en', dest='fr').text.capitalize()
            except Exception: desc_fr = desc_en.capitalize()
            self.last_description = desc_fr
            self.root.after(0, self.animate_text_fade_in, self.description_text, desc_fr)
        except Exception as e: self.root.after(0, self.update_description, f"Erreur de description: {e}", True)

    def detect_objects(self, image):
        try:
            inputs = detr_processor(images=image.convert("RGB"), return_tensors="pt"); outputs = detr_model(**inputs)
            target_sizes = torch.tensor([image.size[::-1]])
            self.last_results = detr_processor.post_process_object_detection(outputs, target_sizes=target_sizes, threshold=CONFIDENCE_THRESHOLD)[0]
            self.root.after(0, self.display_detection_results, self.last_results)
        except Exception as e: print(f"Erreur de détection: {e}"); self.last_results = None

    # --- EXPORT EXCEL ---
    def save_results_to_excel(self):
        if not self.original_image or not self.last_results:
            messagebox.showinfo("Aucune Donnée", "Veuillez d'abord analyser une image avant de sauvegarder.")
            return
        
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Fichiers Excel", "*.xlsx")], title="Sauvegarder le rapport d'analyse")
        if not filepath: return

        try:
            self.update_status("Création du fichier Excel...")
            # 1. Créer le classeur et la feuille
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport d'Analyse"

            ws['A1'] = "Description de la Scène"; ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            ws['A2'] = self.last_description; ws.merge_cells('A2:F2')

            ws['A4'] = "Objets Détectés"; ws['A4'].font = openpyxl.styles.Font(bold=True, size=14)
            ws.append(["", "Objet", "Confiance"])
            ws['B5'].font = openpyxl.styles.Font(bold=True); ws['C5'].font = openpyxl.styles.Font(bold=True)
            
            for i, (score, label, box) in enumerate(zip(self.last_results["scores"], self.last_results["labels"], self.last_results["boxes"])):
                label_en = detr_model.config.id2label[label.item()]
                label_fr = COCO_LABELS_FR.get(label_en, label_en).capitalize()
                ws.append([f"{i+1}.", label_fr, f"{score.item():.1%}"])

            img_with_boxes = self.generate_image_with_detections()
            
            original_img_path = os.path.join(os.path.dirname(filepath), "temp_original.png")
            detected_img_path = os.path.join(os.path.dirname(filepath), "temp_detected.png")
            self.original_image.save(original_img_path)
            img_with_boxes.save(detected_img_path)
            
            ws['E1'] = "Image Originale"; ws['E1'].font = openpyxl.styles.Font(bold=True, size=14)
            img_orig = OpenpyxlImage(original_img_path)
            img_orig.width, img_orig.height = 300, 300 * (img_orig.height / img_orig.width)
            ws.add_image(img_orig, 'E2')

            ws['E15'] = "Image avec Détections"; ws['E15'].font = openpyxl.styles.Font(bold=True, size=14)
            img_det = OpenpyxlImage(detected_img_path)
            img_det.width, img_det.height = 300, 300 * (img_det.height / img_det.width)
            ws.add_image(img_det, 'E16')
            
            ws.column_dimensions['B'].width = 25; ws.column_dimensions['C'].width = 15; ws.column_dimensions['E'].width = 40

            wb.save(filepath)
            os.remove(original_img_path) # Nettoyer les fichiers temporaires
            os.remove(detected_img_path)
            
            self.update_status("Rapport sauvegardé avec succès !")
            messagebox.showinfo("Succès", f"Le rapport a été sauvegardé ici :\n{filepath}")
        except Exception as e:
            self.update_status("Erreur lors de la sauvegarde.")
            messagebox.showerror("Erreur de Sauvegarde", f"Une erreur est survenue :\n{e}")

    def generate_image_with_detections(self):
        img_copy = self.original_image.copy().convert("RGB")
        draw = ImageDraw.Draw(img_copy)
        try: font = ImageFont.truetype("arial.ttf", 20)
        except IOError: font = ImageFont.load_default()

        for score, label, box in zip(self.last_results["scores"], self.last_results["labels"], self.last_results["boxes"]):
            box = box.tolist()
            label_en = detr_model.config.id2label[label.item()]
            label_fr = COCO_LABELS_FR.get(label_en, label_en)
            
            draw.rectangle(box, outline=self.HIGHLIGHT_COLOR, width=3)
            text = f"{label_fr.capitalize()} ({score:.0%})"
            text_bbox = draw.textbbox((box[0], box[1] - 25), text, font=font)
            draw.rectangle(text_bbox, fill=self.HIGHLIGHT_COLOR)
            draw.text((box[0] + 5, box[1] - 25), text, fill="white", font=font)
        return img_copy

    def display_detection_results(self, results):
        self.object_widgets = []
        sorted_results = sorted(zip(results["scores"], results["labels"], results["boxes"]), key=lambda x: x[0], reverse=True)
        for i, (score, label, box) in enumerate(sorted_results):
            label_en = detr_model.config.id2label[label.item()]
            label_fr = COCO_LABELS_FR.get(label_en, label_en).capitalize()
            list_item_frame = RoundedFrame(self.objects_list_frame, radius=10, bg_color=self.ITEM_BG)
            tk.Label(list_item_frame, text=label_fr, font=(self.FONT_FAMILY, 11), bg=self.ITEM_BG, fg=self.TEXT_COLOR, anchor="w", padx=10, pady=8).pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(list_item_frame, text=f"{score.item():.0%}", font=(self.FONT_FAMILY, 11, "bold"), bg=self.ITEM_BG, fg=self.ACCENT_COLOR, anchor="e", padx=10, pady=8).pack(side=tk.RIGHT)
            self.root.after(i * 100, lambda f=list_item_frame: f.pack(fill=tk.X, pady=3))
            box_id = self.canvas.create_rectangle(0,0,0,0, outline=self.ACCENT_COLOR, width=2)
            widget_info = {'list_item_frame': list_item_frame, 'box_id': box_id, 'box_coords': box.tolist(), 'label_fr': label_fr}
            self.object_widgets.append(widget_info)
            list_item_frame.bind("<Enter>", lambda e, idx=i: self.highlight_object(idx))
            list_item_frame.bind("<Leave>", lambda e, idx=i: self.unhighlight_object(idx))
        self.redraw_detection_boxes()
    def display_image(self): self.root.update_idletasks(); self.redraw_canvas_image(); self.animate_image_fade_in()
    def on_resize(self, event=None):
        if hasattr(self, 'resize_job'): self.root.after_cancel(self.resize_job)
        self.resize_job = self.root.after(250, self.redraw_all_canvas_elements)
    def redraw_all_canvas_elements(self):
        if self.original_image: self.redraw_canvas_image(); self.redraw_detection_boxes()
    def redraw_canvas_image(self):
        canvas_w, canvas_h = self.main_frame.winfo_width(), self.main_frame.winfo_height()
        if not self.original_image or canvas_w < 20 or canvas_h < 20: return
        img_copy = self.original_image.copy(); img_copy.thumbnail((canvas_w - 40, canvas_h - 40), Image.Resampling.LANCZOS)
        self.displayed_image_pil = img_copy; self.displayed_image_tk = ImageTk.PhotoImage(self.displayed_image_pil)
        x, y = canvas_w / 2, canvas_h / 2
        if self.image_on_canvas: self.canvas.coords(self.image_on_canvas, x, y); self.canvas.itemconfig(self.image_on_canvas, image=self.displayed_image_tk)
        else: self.image_on_canvas = self.canvas.create_image(x, y, anchor="center", image=self.displayed_image_tk)
    def redraw_detection_boxes(self):
        if not self.displayed_image_pil or not self.object_widgets: return
        img_w, img_h = self.displayed_image_pil.size; orig_w, orig_h = self.original_image.size
        canvas_w, canvas_h = self.main_frame.winfo_width(), self.main_frame.winfo_height()
        offset_x, offset_y = (canvas_w - img_w) / 2, (canvas_h - img_h) / 2
        for info in self.object_widgets:
            x0, y0, x1, y1 = info['box_coords']
            coords = (offset_x + x0*(img_w/orig_w), offset_y + y0*(img_h/orig_h), offset_x + x1*(img_w/orig_w), offset_y + y1*(img_h/orig_h))
            self.canvas.coords(info['box_id'], coords)
    def on_canvas_hover(self, event):
        found_index = None
        for i, info in enumerate(self.object_widgets):
            coords = self.canvas.coords(info['box_id'])
            if coords and coords[0] <= event.x <= coords[2] and coords[1] <= event.y <= coords[3]: found_index = i; break
        if self.current_hovered_index != found_index:
            if self.current_hovered_index is not None: self.unhighlight_object(self.current_hovered_index)
            if found_index is not None: self.highlight_object(found_index)
            self.current_hovered_index = found_index
    def on_canvas_leave(self, event):
        if self.current_hovered_index is not None: self.unhighlight_object(self.current_hovered_index); self.current_hovered_index = None
    def highlight_object(self, index):
        if index >= len(self.object_widgets): return
        info = self.object_widgets[index]
        info['list_item_frame'].bg_color = self.HIGHLIGHT_COLOR; info['list_item_frame']._draw_rounded_corners()
        self.canvas.itemconfig(info['box_id'], outline=self.HIGHLIGHT_COLOR, width=4); self.canvas.tag_raise(info['box_id'])
        x0, y0, _, _ = self.canvas.coords(info['box_id'])
        text_id = self.canvas.create_text(x0, y0 - 10, text=f" {info['label_fr']} ", anchor="sw", font=(self.FONT_FAMILY, 10, "bold"), fill="white", tags="hover_label")
        bg_id = self.canvas.create_rectangle(self.canvas.bbox(text_id), fill="black", outline="", tags="hover_label")
        self.canvas.tag_raise(text_id, bg_id)
    def unhighlight_object(self, index):
        if index >= len(self.object_widgets): return
        info = self.object_widgets[index]
        info['list_item_frame'].bg_color = self.ITEM_BG; info['list_item_frame']._draw_rounded_corners()
        self.canvas.itemconfig(info['box_id'], outline=self.ACCENT_COLOR, width=2)
        self.canvas.delete("hover_label")
    def animate_image_fade_in(self, step=0):
        if step > 25 or not self.displayed_image_pil: return
        alpha = int(255 * (step / 25)); img = self.displayed_image_pil.copy().convert("RGBA"); img.putalpha(alpha)
        self._fade_in_tk_img = ImageTk.PhotoImage(img)
        self.canvas.itemconfig(self.image_on_canvas, image=self._fade_in_tk_img)
        self.root.after(15, self.animate_image_fade_in, step + 1)
    def animate_text_fade_in(self, widget, text, step=0):
        if step > 20: return
        start, end = int(self.CANVAS_BG[1:3], 16), int(self.TEXT_COLOR[1:3], 16)
        val = int(start + (end - start) * (step / 20))
        widget.config(text=text, fg=f'#{val:02x}{val:02x}{val:02x}')
        self.root.after(15, self.animate_text_fade_in, widget, text, step + 1)
    def _animate_panel(self, target_relx, step):
        if step > 25:
            if target_relx == 1.0: self.results_panel.place_forget()
            return
        current_relx = float(self.results_panel.place_info().get('relx', 1.0)); progress = 1 - (1 - step / 25) ** 3
        new_relx = current_relx + (target_relx - current_relx) * progress; self.results_panel.place_configure(relx=new_relx)
        self.root.after(10, self._animate_panel, target_relx, step + 1)
    def clear_previous_analysis(self):
        if hasattr(self, 'resize_job'): self.root.after_cancel(self.resize_job)
        self.top_bar.pack_forget(); self.results_panel.place_forget(); self.description_text.config(text="")
        for widget in self.objects_list_frame.winfo_children(): widget.destroy()
        if self.image_on_canvas: self.canvas.delete("all"); self.image_on_canvas = None
        self.object_widgets = []; self.current_hovered_index = None; self.last_results = None; self.last_description = ""
    def show_welcome_screen(self): self.btn_upload.pack(pady=20); self.loading_label.pack_forget(); self.welcome_frame.place(relx=0.5, rely=0.5, anchor="center")
    def show_loading_screen(self, message): self.update_status(message); self.btn_upload.pack_forget(); self.loading_label.config(text=message); self.loading_label.pack(pady=20); self.welcome_frame.place(relx=0.5, rely=0.5, anchor="center")
    def hide_welcome_screen(self): self.welcome_frame.place_forget()
    def animate_results_panel_in(self): self.results_panel.place(relx=1.0, rely=0, relwidth=0.32, relheight=1); self._animate_panel(target_relx=(1.0 - 0.32), step=0)
    def update_status(self, message): self.status_label.config(text=message)
    def update_description(self, msg, is_error=False): self.description_text.config(text=msg, fg=self.HIGHLIGHT_COLOR if is_error else self.TEXT_COLOR)



root = tk.Tk()
app = FullImageAnalyzerApp(root)
threading.Thread(target=app.load_models_in_thread, daemon=True).start()
root.mainloop()
